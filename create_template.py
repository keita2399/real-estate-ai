"""
重要事項説明書 Excelテンプレート生成スクリプト（本格版）
実行: python create_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ===== スタイル定義 =====
def border(top="thin", bottom="thin", left="thin", right="thin", color="000000"):
    s = lambda w, c=color: Side(style=w, color=c) if w else Side(style=None)
    return Border(top=s(top), bottom=s(bottom), left=s(left), right=s(right))

THICK = border("medium","medium","medium","medium")
THIN  = border()
BTOP  = border(top="medium", bottom="thin", left="medium", right="medium")
BBOX  = border(top="thin",   bottom="medium", left="medium", right="medium")
BLEFT = border(top="thin",   bottom="thin",   left="medium", right="thin")
BRIGHT= border(top="thin",   bottom="thin",   left="thin",   right="medium")

def fill(hex): return PatternFill("solid", fgColor=hex)
def font(name="MS Gothic", size=9, bold=False, color="000000"):
    return Font(name=name, size=size, bold=bold, color=color)
def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# カラーパレット
C_TITLE  = "1F3864"   # 濃紺
C_SEC    = "2E75B6"   # 青
C_SUBHD  = "BDD7EE"   # 薄青
C_LABEL  = "DEEAF1"   # 薄水色
C_WHITE  = "FFFFFF"
C_YELLOW = "FFFFC0"   # 薄黄（入力欄）
C_GRAY   = "F2F2F2"

def set(ws, ref, val="", f=None, fi=None, al=None, bo=None):
    c = ws[ref]
    c.value     = val
    if f:  c.font      = f
    if fi: c.fill      = fi
    if al: c.alignment = al
    if bo: c.border    = bo

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def row_height(ws, row, h): ws.row_dimensions[row].height = h
def col_width(ws, col, w):  ws.column_dimensions[get_column_letter(col)].width = w

# ===== セクションヘッダー描画 =====
def section_header(ws, row, title, c1=1, c2=12):
    merge(ws, row, c1, row, c2)
    cell = ws.cell(row, c1)
    cell.value     = title
    cell.font      = Font(name="MS Gothic", size=10, bold=True, color=C_WHITE)
    cell.fill      = fill(C_SEC)
    cell.alignment = align("left")
    cell.border    = THICK
    row_height(ws, row, 20)

# ===== 2列ラベル行 =====
def label_row(ws, row, label, col_label=2, col_val=4, col_end=12,
              fill_label=C_LABEL, fill_val=C_YELLOW, h=18):
    # ラベルセル
    merge(ws, row, col_label, row, col_label+1)
    c = ws.cell(row, col_label)
    c.value = label; c.font = font(bold=True); c.fill = fill(fill_label)
    c.alignment = align("left"); c.border = THIN
    # 値セル
    merge(ws, row, col_val, row, col_end)
    c = ws.cell(row, col_val)
    c.value = ""; c.fill = fill(fill_val)
    c.alignment = align("left"); c.border = THIN
    row_height(ws, row, h)

# ===== 横並び2項目行 =====
def two_col_row(ws, row, label1, label2, col_end=12,
                fill_label=C_LABEL, fill_val=C_YELLOW):
    mid = (col_end + 2) // 2
    # 左ラベル
    merge(ws, row, 2, row, 3)
    c = ws.cell(row, 2); c.value = label1
    c.font = font(bold=True); c.fill = fill(fill_label)
    c.alignment = align("left"); c.border = THIN
    # 左値
    merge(ws, row, 4, row, mid-1)
    c = ws.cell(row, 4); c.fill = fill(fill_val)
    c.alignment = align("left"); c.border = THIN
    # 右ラベル
    merge(ws, row, mid, row, mid+1)
    c = ws.cell(row, mid); c.value = label2
    c.font = font(bold=True); c.fill = fill(fill_label)
    c.alignment = align("left"); c.border = THIN
    # 右値
    merge(ws, row, mid+2, row, col_end)
    c = ws.cell(row, mid+2); c.fill = fill(fill_val)
    c.alignment = align("left"); c.border = THIN
    row_height(ws, row, 18)


def create_template(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "重要事項説明書"

    # ----- 列幅 -----
    widths = {1:2, 2:6, 3:10, 4:10, 5:10, 6:10,
              7:10, 8:10, 9:10, 10:10, 11:10, 12:2}
    for c, w in widths.items(): col_width(ws, c, w)

    # ----- 印刷設定 -----
    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage  = True

    r = 1  # 現在行カーソル

    # =====================================================
    # タイトル
    # =====================================================
    row_height(ws, r, 8)
    r += 1

    merge(ws, r, 1, r+1, 12)
    c = ws.cell(r, 1)
    c.value = "重　要　事　項　説　明　書"
    c.font  = Font(name="MS Gothic", size=18, bold=True, color=C_TITLE)
    c.fill  = fill("DEEAF1")
    c.alignment = align()
    c.border = THICK
    row_height(ws, r, 30); row_height(ws, r+1, 30)
    r += 2

    merge(ws, r, 1, r, 12)
    c = ws.cell(r, 1)
    c.value = "（土地・建物の売買用）　　　　　　宅地建物取引業法第35条の規定に基づき、下記のとおり説明します。"
    c.font  = Font(name="MS Gothic", size=8, color="444444")
    c.fill  = fill("DEEAF1")
    c.alignment = align("left")
    c.border = border(top="thin", bottom="medium", left="medium", right="medium")
    row_height(ws, r, 16)
    r += 1

    row_height(ws, r, 6); r += 1

    # =====================================================
    # 宅建業者・宅建士情報
    # =====================================================
    section_header(ws, r, "■ 宅地建物取引業者・宅地建物取引士"); r += 1

    for label in ["商号・名称", "免許証番号", "所在地", "電話番号", "宅建士氏名", "登録番号"]:
        label_row(ws, r, label, fill_val=C_WHITE); r += 1

    row_height(ws, r, 6); r += 1

    # 日付・買主欄
    merge(ws, r, 2, r, 5)
    ws.cell(r, 2).value = "説明年月日"
    ws.cell(r, 2).font  = font(bold=True); ws.cell(r, 2).fill = fill(C_LABEL)
    ws.cell(r, 2).alignment = align("left"); ws.cell(r, 2).border = THIN
    merge(ws, r, 6, r, 12)
    ws.cell(r, 6).fill = fill(C_YELLOW); ws.cell(r, 6).border = THIN
    row_height(ws, r, 18); r += 1

    merge(ws, r, 2, r, 5)
    ws.cell(r, 2).value = "買主（説明を受けた者）"
    ws.cell(r, 2).font  = font(bold=True); ws.cell(r, 2).fill = fill(C_LABEL)
    ws.cell(r, 2).alignment = align("left"); ws.cell(r, 2).border = THIN
    merge(ws, r, 6, r, 12)
    ws.cell(r, 6).fill = fill(C_YELLOW); ws.cell(r, 6).border = THIN
    row_height(ws, r, 18); r += 1

    row_height(ws, r, 6); r += 1

    # =====================================================
    # 第1 物件の表示
    # =====================================================
    section_header(ws, r, "第１　取引する宅地又は建物の表示"); r += 1

    merge(ws, r, 2, r, 12)
    ws.cell(r, 2).value = "【土 地】"
    ws.cell(r, 2).font  = Font(name="MS Gothic", size=9, bold=True, color=C_SEC)
    ws.cell(r, 2).fill  = fill(C_SUBHD)
    ws.cell(r, 2).alignment = align("left")
    ws.cell(r, 2).border = THIN
    row_height(ws, r, 16); r += 1

    for label in ["所　在", "地　番", "地　目"]:
        label_row(ws, r, label); r += 1
    two_col_row(ws, r, "地　積", "（㎡）"); r += 1

    merge(ws, r, 2, r, 12)
    ws.cell(r, 2).value = "【建 物】"
    ws.cell(r, 2).font  = Font(name="MS Gothic", size=9, bold=True, color=C_SEC)
    ws.cell(r, 2).fill  = fill(C_SUBHD)
    ws.cell(r, 2).alignment = align("left")
    ws.cell(r, 2).border = THIN
    row_height(ws, r, 16); r += 1

    for label in ["所　在", "家屋番号", "種　類", "構　造"]:
        label_row(ws, r, label); r += 1
    two_col_row(ws, r, "床面積", "（㎡）"); r += 1
    label_row(ws, r, "建築年月"); r += 1

    row_height(ws, r, 6); r += 1

    # =====================================================
    # 第2 登記記録（所有権）
    # =====================================================
    section_header(ws, r, "第２　登記記録に記録された事項（甲区：所有権に関する事項）"); r += 1

    for label in ["所有者氏名", "所有者住所", "取得原因", "取得日", "共有持分"]:
        label_row(ws, r, label); r += 1

    row_height(ws, r, 6); r += 1

    # =====================================================
    # 第3 登記記録（抵当権）
    # =====================================================
    section_header(ws, r, "第３　登記記録に記録された事項（乙区：所有権以外の権利に関する事項）"); r += 1

    two_col_row(ws, r, "抵当権・根抵当権", "有　無"); r += 1
    two_col_row(ws, r, "債権額（極度額）", "設定日"); r += 1
    label_row(ws, r, "抵当権者（債権者）"); r += 1
    label_row(ws, r, "債　務　者"); r += 1

    row_height(ws, r, 6); r += 1

    # =====================================================
    # 第4 都市計画
    # =====================================================
    section_header(ws, r, "第４　都市計画法・建築基準法等の法令に基づく制限の概要"); r += 1

    label_row(ws, r, "用途地域"); r += 1
    two_col_row(ws, r, "建ぺい率（%）", "容積率（%）"); r += 1
    label_row(ws, r, "防火・準防火地域"); r += 1
    label_row(ws, r, "高度地区・斜線制限"); r += 1

    row_height(ws, r, 6); r += 1

    # =====================================================
    # 第5 インフラ
    # =====================================================
    section_header(ws, r, "第５　飲用水・電気・ガスの供給施設及び排水施設の整備状況"); r += 1

    two_col_row(ws, r, "飲用水（上水道）", "排水（下水道）"); r += 1
    two_col_row(ws, r, "ガス",             "電気"); r += 1

    row_height(ws, r, 6); r += 1

    # =====================================================
    # 第6 ハザードマップ
    # =====================================================
    section_header(ws, r, "第６　水害ハザードマップにおける対象物件の所在地（宅建業法施行規則第16条の4の3第3号の2）"); r += 1

    for label in ["洪水浸水想定区域", "雨水出水浸水想定区域", "高潮浸水想定区域",
                  "土砂災害警戒区域", "津波災害警戒区域", "液状化リスク", "最寄り指定避難場所"]:
        label_row(ws, r, label); r += 1

    row_height(ws, r, 6); r += 1

    # =====================================================
    # 注記
    # =====================================================
    merge(ws, r, 1, r+1, 12)
    c = ws.cell(r, 1)
    c.value = (
        "【注意事項】　本書面はAI（Gemini 2.5 Flash）による情報抽出を基に自動生成されています。"
        "記載内容については宅地建物取引士が必ず確認・修正を行い、最終的な責任のもとで使用してください。"
        "　　Powered by Claude Code × Gemini API"
    )
    c.font  = Font(name="MS Gothic", size=7, italic=True, color="666666")
    c.fill  = fill("FFF9C4")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = THICK
    row_height(ws, r, 14); row_height(ws, r+1, 14)
    r += 2

    row_height(ws, r, 8)

    # 印刷範囲
    ws.print_area = f"A1:{get_column_letter(12)}{r}"

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"テンプレート生成完了: {output_path}")


if __name__ == "__main__":
    create_template("template/jyuujiku.xlsx")
