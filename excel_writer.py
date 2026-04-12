"""
抽出したデータをExcelテンプレートに書き込むモジュール
"""
import logging
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from config import EXCEL_CELL_MAPPING

logger = logging.getLogger(__name__)

# セルマッピング（config.pyとテンプレートの対応）
# config.pyのキーとExcelのセルを明示的にマッピング
FIELD_TO_CELL = {
    # 物件情報（土地）
    "所在":             "D19",
    "地番":             "D20",
    "地目":             "D21",
    "地積":             "D22",
    # 物件情報（建物）
    "建物所在":         "D24",
    "家屋番号":         "D25",
    "種類":             "D26",
    "構造":             "D27",
    "床面積":           "D28",
    "建築年月":         "D29",
    # 所有権
    "所有者氏名":       "D32",
    "所有者住所":       "D33",
    "取得原因":         "D34",
    "取得日":           "D35",
    "共有持分":         "D36",
    # 抵当権（two_col: 左=D, 右=I）
    "抵当権有無":       "I39",
    "債権額":           "D40",
    "設定日":           "I40",
    "債権者":           "D41",
    "債務者":           "D42",
    # 都市計画（two_col: 建ぺい率=D46, 容積率=I46）
    "用途地域":         "D45",
    "建ぺい率":         "D46",
    "容積率":           "I46",
    "防火地域":         "D47",
    "高度地区":         "D48",
    # インフラ（two_col: 上水道=D51, 下水道=I51, ガス=D52, 電気=I52）
    "上水道":           "D51",
    "下水道":           "I51",
    "ガス":             "D52",
    "電気":             "I52",
    # ハザードマップ
    "洪水浸水想定区域": "D55",
    "土砂災害警戒区域": "D58",
    "津波災害警戒区域": "D59",
    "液状化リスク":     "D60",
    "避難場所":         "D61",
}

VALUE_FONT  = Font(name="メイリオ", size=9, color="1F1F1F")
FILLED_FILL = PatternFill("solid", fgColor="E2EFDA")  # 薄い緑＝AI入力済み


COMPANY_CELL_MAPPING = {
    "商号名称":   "D7",
    "免許証番号": "D8",
    "所在地":     "D9",
    "電話番号":   "D10",
    "宅建士氏名": "D11",
    "登録番号":   "D12",
}

COMPANY_FONT  = Font(name="MS Gothic", size=9, color="1F1F1F")
COMPANY_FILL  = PatternFill("solid", fgColor="DEEAF1")


def write_to_excel(
    extracted: dict,
    template_path: str,
    output_dir: str = "output",
    prefix: str = "",
    company_info: dict | None = None,
) -> str:
    """
    抽出データをExcelテンプレートに書き込んで保存する。
    戻り値: 保存したファイルパス
    """
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {template_path}")

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = prefix or "重要事項説明書"
    output_path = Path(output_dir) / f"{name}_{timestamp}.xlsx"

    shutil.copy(template, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["重要事項説明書"]

    filled = 0
    skipped = 0

    for field, cell_ref in FIELD_TO_CELL.items():
        value = extracted.get(field, "")
        if not value or value in ("不明", "記載なし", "なし"):
            skipped += 1
            continue
        try:
            # 結合セルの先頭セルを特定して書き込む
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_str, row_num = coordinate_from_string(cell_ref)
            col_num = column_index_from_string(col_str)
            target_row, target_col = row_num, col_num
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row_num <= merged_range.max_row and
                        merged_range.min_col <= col_num <= merged_range.max_col):
                    target_row = merged_range.min_row
                    target_col = merged_range.min_col
                    break
            cell = ws.cell(target_row, target_col)
            cell.value = value
            cell.font  = VALUE_FONT
            cell.fill  = FILLED_FILL
            filled += 1
        except Exception as e:
            logger.warning(f"セル {cell_ref} ({field}) への書き込みに失敗: {e}")

    # 会社情報を書き込む
    if company_info:
        for field, cell_ref in COMPANY_CELL_MAPPING.items():
            value = company_info.get(field, "")
            if not value:
                continue
            try:
                from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                col_str, row_num = coordinate_from_string(cell_ref)
                col_num = column_index_from_string(col_str)
                target_row, target_col = row_num, col_num
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row_num <= merged_range.max_row and
                            merged_range.min_col <= col_num <= merged_range.max_col):
                        target_row = merged_range.min_row
                        target_col = merged_range.min_col
                        break
                cell = ws.cell(target_row, target_col)
                cell.value = value
                cell.font  = COMPANY_FONT
                cell.fill  = COMPANY_FILL
            except Exception as e:
                logger.warning(f"会社情報セル {cell_ref} ({field}) への書き込みに失敗: {e}")

    # 説明年月日を自動入力
    from datetime import date
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    today = date.today()
    _date_str = f"令和{today.year - 2018}年{today.month}月{today.day}日"
    _row, _col = 14, 6  # F14 が説明年月日の値セル先頭
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= 14 <= mr.max_row and mr.min_col <= 6 <= mr.max_col:
            _row, _col = mr.min_row, mr.min_col
            break
    _cell = ws.cell(_row, _col)
    _cell.value = _date_str
    _cell.font  = COMPANY_FONT
    _cell.fill  = COMPANY_FILL

    wb.save(output_path)
    logger.info(f"Excel保存完了: {output_path}  ({filled}件記入 / {skipped}件スキップ)")
    return str(output_path)


def preview_extracted(extracted: dict):
    """抽出結果をターミナルに整形表示する"""
    section_map = {
        "物件情報（土地）":    ["所在", "地番", "地目", "地積"],
        "物件情報（建物）":    ["建物所在", "家屋番号", "種類", "構造", "床面積", "建築年月"],
        "所有権":              ["所有者氏名", "所有者住所", "取得原因", "取得日", "共有持分"],
        "抵当権等":            ["抵当権有無", "債権額", "債権者", "債務者", "設定日"],
        "都市計画":            ["用途地域", "建ぺい率", "容積率", "防火地域", "高度地区"],
        "インフラ":            ["上水道", "下水道", "ガス", "電気"],
        "ハザードマップ":      ["洪水浸水想定区域", "土砂災害警戒区域", "津波災害警戒区域", "液状化リスク", "避難場所"],
    }
    print("\n" + "=" * 60)
    print("  Gemini Extraction Preview")
    print("=" * 60)
    for section, fields in section_map.items():
        print(f"\n[{section}]")
        for f in fields:
            val = extracted.get(f, "（なし）")
            print(f"   {f:<18} : {val}")
    print("=" * 60)
